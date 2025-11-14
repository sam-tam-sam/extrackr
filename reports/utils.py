from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Sum
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration


def generate_pdf_report(transactions, report_type, user):
    """Generate PDF report"""
    
    # Calculate summary data
    income_total = transactions.filter(transaction_type='income').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    expense_total = transactions.filter(transaction_type='expense').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    net_balance = income_total - expense_total
    
    # Get date range
    if transactions.exists():
        date_range = {
            'from': transactions.order_by('date').first().date,
            'to': transactions.order_by('date').last().date
        }
    else:
        date_range = {
            'from': timezone.now().date(),
            'to': timezone.now().date()
        }
    
    # Prepare context
    context = {
        'user': user,
        'transactions': transactions,
        'report_type': report_type,
        'date_range': date_range,
        'generated_date': timezone.now(),
        'summary': {
            'income': income_total,
            'expenses': expense_total,
            'net_balance': net_balance,
            'total_transactions': transactions.count()
        }
    }
    
    # Render HTML template
    html_string = render_to_string('reports/pdf_template.html', context)
    
    # Generate PDF
    font_config = FontConfiguration()
    html = HTML(string=html_string)
    pdf_file = html.write_pdf(font_config=font_config)
    
    # Create response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"extrackr_report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def generate_excel_report(transactions, report_type, user):
    """Generate Excel report"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_font = Font(bold=True)
    income_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    expense_fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Header information
    ws['A1'] = f"extrackr Financial Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"User: {user.get_full_name() or user.username}"
    ws['A3'] = f"Report Type: {report_type.title()}"
    ws['A4'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Add empty row
    ws.append([])
    
    # Summary section
    income_total = transactions.filter(transaction_type='income').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    expense_total = transactions.filter(transaction_type='expense').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    net_balance = income_total - expense_total
    
    summary_start_row = ws.max_row + 1
    ws[f'A{summary_start_row}'] = "Summary"
    ws[f'A{summary_start_row}'].font = Font(size=14, bold=True)
    
    ws[f'A{summary_start_row + 1}'] = "Total Income:"
    ws[f'B{summary_start_row + 1}'] = float(income_total)
    ws[f'B{summary_start_row + 1}'].number_format = '"$"#,##0.00'
    
    ws[f'A{summary_start_row + 2}'] = "Total Expenses:"
    ws[f'B{summary_start_row + 2}'] = float(expense_total)
    ws[f'B{summary_start_row + 2}'].number_format = '"$"#,##0.00'
    
    ws[f'A{summary_start_row + 3}'] = "Net Balance:"
    ws[f'B{summary_start_row + 3}'] = float(net_balance)
    ws[f'B{summary_start_row + 3}'].number_format = '"$"#,##0.00'
    
    # Add empty row
    ws.append([])
    
    # Transactions header
    header_row = ws.max_row + 1
    headers = ['Date', 'Type', 'Category', 'Description', 'Amount', 'Balance']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Add transactions
    current_row = header_row + 1
    running_balance = 0
    
    for transaction in transactions.order_by('date'):
        ws.cell(row=current_row, column=1, value=transaction.date)
        ws.cell(row=current_row, column=1).number_format = 'YYYY-MM-DD'
        
        ws.cell(row=current_row, column=2, value=transaction.get_transaction_type_display())
        ws.cell(row=current_row, column=3, value=transaction.category.name)
        ws.cell(row=current_row, column=4, value=transaction.description or '')
        
        amount = float(transaction.amount)
        ws.cell(row=current_row, column=5, value=amount)
        ws.cell(row=current_row, column=5).number_format = '"$"#,##0.00'
        
        # Update running balance
        if transaction.transaction_type == 'income':
            running_balance += amount
            ws.cell(row=current_row, column=5).fill = income_fill
        else:
            running_balance -= amount
            ws.cell(row=current_row, column=5).fill = expense_fill
        
        ws.cell(row=current_row, column=6, value=running_balance)
        ws.cell(row=current_row, column=6).number_format = '"$"#,##0.00'
        
        current_row += 1
    
    # Add totals row
    totals_row = current_row
    ws.cell(row=totals_row, column=4, value="TOTALS")
    ws.cell(row=totals_row, column=4).font = total_font
    
    # Income total
    ws.cell(row=totals_row, column=5, value=float(income_total))
    ws.cell(row=totals_row, column=5).number_format = '"$"#,##0.00'
    ws.cell(row=totals_row, column=5).font = total_font
    ws.cell(row=totals_row, column=5).fill = income_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"extrackr_report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response